import pandas as pd
import pytz
from datetime import time
import json
from collections import defaultdict

fil = open("odsherred/wcif.json")
data = json.load(fil)

timezone = pytz.timezone(data["schedule"]["venues"][0]["timezone"])
# Convert to DWs names
mapping = {'222':'2x2','333':'3x3','444':'4x4','555':'5x5','666':'6x6','777':'7x7','sq1':'Square-1'
,'skewb':'Skewb','minx':'Megaminx', 'clock':'Clock','pyram':'Pyraminx','333oh':'3x3 OH','333fm':'3x3 FMC',
'333bf':'3x3 BLD','444bf':'4x4 BLD','555bf':'555 BLD','333mbf':'3x3 MBLD',
'tutorial':'Deltagarintroduktion','lunch':'Lunch','misc':'Städning','awards':'Prisutdelning'}

eventcount = defaultdict(int) # The amount of competitors per event

for person in data['persons']:
	try:
		if person['registration']['status'] == 'accepted':
			for event in person['registration']['eventIds']:
				eventcount[mapping[event]]+=1
				# people[person['name']].append(event)
	except TypeError:
		pass

# print(eventcount
eventCut = defaultdict(dict)
eventPro = defaultdict(dict)

for event in data["events"]:
	for round in event['rounds']:
		eventname = round['id'].split('-')
		if round['timeLimit']['cumulativeRoundIds']:
			eventCut[mapping[eventname[0]]][int(eventname[1][1:])] = (1,round['timeLimit']['centiseconds']*6000)
		elif round['cutoff']:
			eventCut[mapping[eventname[0]]][int(eventname[1][1:])] = (2,round['cutoff']*6000)
		else:
			eventCut[mapping[eventname[0]]][int(eventname[1][1:])] = (0,0)

		if round['advancementCondition']:
			if round['advancementCondition']['type'] == 'ranking':
				eventPro[mapping[eventname[0]]][int(eventname[1][1:])+1] = (1,round['advancementCondition']['level'])
			elif round['advancementCondition']['type'] == 'percent':
				eventPro[mapping[eventname[0]]][int(eventname[1][1:])+1]  = (0,round['advancementCondition']['level'])
			else:
				eventPro[mapping[eventname[0]]][int(eventname[1][1:])+1]  = (0,75)
		# eventCut[round['id'].split('-')[1]] = 

schedule = [] 
for room in data["schedule"]["venues"][0]['rooms']:
	for val in room["activities"]:
		starttime = pd.Timestamp(val['startTime'][:-1]).tz_localize(pytz.utc).tz_convert(timezone)
		endtime = pd.Timestamp(val['endTime'][:-1]).tz_localize(pytz.utc).tz_convert(timezone)
		# print(val['activityCode'],val.keys())
		ev = val['activityCode'].split('-')
		if ev[0][0] != 'o':
			ev = mapping[ev[0]] # Map event name
		else:
			ev=mapping[ev[1]] # Not an event or commonly known activity, give it something of a potentially appropriate time
			# print(val)

		schedule.append((ev,starttime,endtime))
schedule.sort(key=lambda x:x[1]) # Sort based on start time

daySplit=[] # see when the schedule devides in two days. It doesn't work for more than 2 days.
def getDaySplit():
	for i in range(1,len(schedule)):
		if schedule[i][1].day == schedule[i-1][1].day:
			pass
		else:
			daySplit.append(i)
getDaySplit()

from openpyxl import load_workbook
wb = load_workbook(filename = 'dwschedule.xlsx') # open the schedule template

sheet_names = wb.sheetnames
sch = wb[sheet_names[0]] # The interesting tab

# start time
sch.cell(row=6,column=1,value=schedule[0][1].hour)
sch.cell(row=6,column=2).value = schedule[0][1].minute

i = 8 # row of first activity
event_hit = defaultdict(int)

if daySplit:
	for val in schedule[:daySplit[0]]:
		event_hit[val[0]] += 1
		sch.cell(row=i,column=3).value = val[0] # set event name
		if val[0] in eventcount: # if it is an event, where we need to add competitors
			if event_hit[val[0]] == 1: # first round
				sch.cell(row=i,column=10).value = eventcount[val[0]] # set competitors
			else: # advancement conditions for the rest
				if eventPro[val[0]][event_hit[val[0]]][0]: # Ranking based
					sch.cell(row=i,column=10).value = eventPro[val[0]][event_hit[val[0]]][1]
				else: # Percentage based
					sch.cell(row=i,column=10).value = int((eventPro[val[0]][event_hit[val[0]]][1]/100)*eventcount[val[0]])

			if eventCut[val[0]][event_hit[val[0]]] == 1: # Cumulative time limit
				sch.cell(row=i,column=8).value = eventCut[val[0]][event_hit[val[0]]][1]
			elif eventCut[val[0]][event_hit[val[0]]] == 2: # Cutoff
				sch.cell(row=i,column=9).value = eventCut[val[0]][event_hit[val[0]]][1]
		i+=1
	sch.cell(row=29,column=1,value=schedule[daySplit[0]][1].hour)
	sch.cell(row=29,column=2).value = schedule[daySplit[0]][1].minute
	i = 31 # jump to day 
	# The rest of the code is basically the same, but too lazy to make pretty non repetitive code
	for val in schedule[daySplit[0]:]:
		event_hit[val[0]] += 1
		sch.cell(row=i,column=3).value = val[0]
		if val[0] in eventcount:
			if event_hit[val[0]] == 1:
				sch.cell(row=i,column=10).value = eventcount[val[0]]
			else:
				if eventPro[val[0]][event_hit[val[0]]][0]:
					eventPro[val[0]][event_hit[val[0]]][1]
				else:
					sch.cell(row=i,column=10).value = int((eventPro[val[0]][event_hit[val[0]]][1]/100)*eventcount[val[0]])
			
			if eventCut[val[0]][event_hit[val[0]]] == 1:
				sch.cell(row=i,column=8).value = eventCut[val[0]][event_hit[val[0]]][1]
			elif eventCut[val[0]][event_hit[val[0]]] == 2:
				sch.cell(row=i,column=9).value = eventCut[val[0]][event_hit[val[0]]][1]
		i+=1
else: # One day comp, again basically the same
	for val in schedule:
		event_hit[val[0]] += 1
		sch.cell(row=i,column=3).value = val[0]
		if val[0] in eventcount:
			if event_hit[val[0]] == 1:
				sch.cell(row=i,column=10).value = eventcount[val[0]]
			else:
				if eventPro[val[0]][event_hit[val[0]]][0]:
					eventPro[val[0]][event_hit[val[0]]][1]
				else:
					sch.cell(row=i,column=10).value = int((eventPro[val[0]][event_hit[val[0]]][1]/100)*eventcount[val[0]])
			
			if eventCut[val[0]][event_hit[val[0]]] == 1:
				sch.cell(row=i,column=8).value = eventCut[val[0]][event_hit[val[0]]][1]
			elif eventCut[val[0]][event_hit[val[0]]] == 2:
				sch.cell(row=i,column=9).value = eventCut[val[0]][event_hit[val[0]]][1]
		i+=1

wb.save('dw4.xlsx') # saving the excel sheet


